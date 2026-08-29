from models import db, Tournament as TournamentModel, Group as GroupModel, Match as MatchModel, Bracket as BracketModel, \
    Player as PlayerModel, PlayerStats as PlayerStatsModel
from player import PlayerHelper
from groupmanager import GroupManager
from playoff import Playoff
import io
import openpyxl


class WebManager:
    """
    Presenter vrstva: Zodpovídá za čtení dat z doménové logiky a jejich 
    transformaci do slovníků a formátů přichystaných pro HTML šablony.
    """

    def __init__(self, tournament_id: int):
        self.tournament_id = tournament_id
        self.tournament = TournamentModel.query.get_or_404(tournament_id)
        self.group_manager = GroupManager(tournament_id, self.tournament.group_match_format)

    def get_groups_page_data(self) -> dict:
        """Připraví data pro stránku se základními skupinami (groups.html)."""
        group_data = {}
        groups = GroupModel.query.filter_by(tournament_id=self.tournament_id, is_consolation=False).order_by(
            GroupModel.name.asc()).all()

        for group in groups:
            ranked_players = self.group_manager.rank_players(group.id, "Group")
            group_data[group.name.replace("Skupina ", "")] = {
                "players": self._format_players(ranked_players, "Group"),
                "matches": self._format_matches(group.matches)
            }
        return group_data

    def get_minigroup_page_data(self) -> dict:
        """Připraví data pro minitabulku útěchy (consolation_minigroup.html)."""
        group_data = {}
        cons_bracket = BracketModel.query.filter_by(
            tournament_id=self.tournament_id,
            is_consolation=True,
            bracket_type="round_robin"
        ).first()

        if cons_bracket:
            # 1. Vytáhneme reálné zápasy minitabulky z DB
            matches = MatchModel.query.filter_by(bracket_id=cons_bracket.id).all()

            # 2. Sesbíráme hráče, kteří už v minitabulce mají zápasy, NEBO už mají group_seed odpovídající vyřazeným
            match_player_ids = {m.player_a_id for m in matches if m.player_a_id} | {m.player_b_id for m in matches if
                                                                                    m.player_b_id}

            # Navíc se podíváme do tree_data, jaké seedy (např "3A", "3B") tato minitabulka vůbec očekává
            tree_data = cons_bracket.tree_data or {}
            expected_seeds = set()
            for slot_a, slot_b in tree_data.get("matches", []):
                if isinstance(slot_a, str): expected_seeds.add(slot_a)
                if isinstance(slot_b, str): expected_seeds.add(slot_b)

            # Najdeme hráče, kteří odpovídají těmto seedům a už mají přiřazené ID / dohráli skupinu
            seeded_players = PlayerModel.query.filter(
                PlayerModel.tournament_id == self.tournament_id,
                PlayerModel.group_seed.in_(expected_seeds)
            ).all()

            all_player_ids = match_player_ids | {p.id for p in seeded_players}
            players = PlayerModel.query.filter(PlayerModel.id.in_(all_player_ids)).all()

            ranked = sorted(
                players,
                key=lambda p: (
                    PlayerHelper.get_or_create_stats(p.id, "minigroup").points,
                    PlayerHelper.difference_of_score(p.id, "minigroup")["Balls"]
                ),
                reverse=True
            )

            group_data[cons_bracket.name] = {
                "players": self._format_players(ranked, "minigroup"),
                "matches": self._format_matches(matches)
            }

        return group_data

    def get_playoff_page_data(self, is_consolation: bool = False) -> dict:
        """Připraví data pro hlavní nebo útěchový pavouk."""
        bracket, rank_offset, stage_name = self._get_playoff_bracket_info(is_consolation)
        p_data = None
        if bracket:
            engine = Playoff(
                tournament_id=self.tournament_id,
                match_format=self.tournament.playoff_match_format,
                stage_name=stage_name,
                playoff_elimination_action=self.tournament.playoff_elimination_action,
                bracket_id=bracket.id,
                rank_offset=rank_offset
            )
            p_data = engine.get_ui_data()
        return p_data

    def handle_playoff_completion(self, is_consolation: bool = False) -> None:
        """Spustí playoff motor pro posun hráčů a zápis výsledků."""
        bracket, rank_offset, stage_name = self._get_playoff_bracket_info(is_consolation)
        if bracket:
            engine = Playoff(
                tournament_id=self.tournament_id,
                match_format=self.tournament.playoff_match_format,
                stage_name=stage_name,
                playoff_elimination_action=self.tournament.playoff_elimination_action,
                bracket_id=bracket.id,
                rank_offset=rank_offset
            )
            engine.check_and_proceed()
            engine.save_to_db()

    def _get_playoff_bracket_info(self, is_consolation: bool):
        """Pomocná metoda pro zjištění parametrů playoff bracketu."""
        if is_consolation:
            bracket = BracketModel.query.filter_by(
                tournament_id=self.tournament_id,
                is_consolation=True,
                bracket_type="elimination"
            ).first()
            main_groups_count = GroupModel.query.filter_by(tournament_id=self.tournament_id,
                                                           is_consolation=False).count()
            rank_offset = self.tournament.advance_per_group * main_groups_count
            stage_name = "consolation"
        else:
            bracket = BracketModel.query.filter_by(
                tournament_id=self.tournament_id,
                name="Hlavní Playoff"
            ).first()
            rank_offset = 0
            stage_name = "main"
        return bracket, rank_offset, stage_name

    # ==========================================
    # PRIVÁTNÍ POMOCNÉ METODY PRO FORMÁTOVÁNÍ
    # ==========================================

    @staticmethod
    def _format_players(players: list, stage_name: str) -> list:
        if not players:
            return []

        # 1. Získáme ID všech hráčů najednou
        player_ids = [p.id for p in players]

        # 2. Vytáhneme statistiky pro VŠECHNY tyto hráče jediným SQL dotazem
        from models import PlayerStats
        all_stats = PlayerStats.query.filter(
            PlayerStats.player_id.in_(player_ids),
            PlayerStats.stage_name == stage_name
        ).all()

        # Uložíme do slovníku pro okamžitý přístup O(1)
        stats_map = {s.player_id: s for s in all_stats}

        ui_data = []
        for p in players:
            stats = stats_map.get(p.id)
            if stats:
                games_win = stats.games_win
                games_lost = stats.games_lost
                balls_diff = stats.balls_win - stats.balls_lost
                points = stats.points
            else:
                games_win, games_lost, balls_diff, points = 0, 0, 0, 0

            ui_data.append({
                "name": p.name,
                "games_win": games_win,
                "games_lost": games_lost,
                "balls_diff": balls_diff,
                "points": points
            })
        return ui_data

    @staticmethod
    def _format_matches(matches: list) -> list:
        ui_data = []
        for m in sorted(matches, key=lambda x: x.id):
            played_sets = []
            if hasattr(m, 'sets') and m.sets:
                # Seřadíme je podle čísla setu, aby šly popořadě (1. set, 2. set...)
                sorted_sets = sorted(m.sets, key=lambda s: s.set_number)
                for s in sorted_sets:
                    played_sets.append((s.score_a, s.score_b))
            ui_data.append({
                "match_id": m.id,
                "player_a_name": m.player_a.name if m.player_a else "TBD",
                "player_b_name": m.player_b.name if m.player_b else "TBD",
                "is_finished": m.is_finished,
                "is_in_progress": getattr(m, "is_in_progress", False),
                "match_format": m.match_format,
                "played_sets": played_sets
            })
        return ui_data

    def generate_results_excel(self) -> io.BytesIO:
        """Vygeneruje Excel soubor s konečným pořadím turnaje a vrátí ho jako BytesIO stream."""
        results_data = db.session.query(PlayerModel, PlayerStatsModel.final_rank) \
            .join(PlayerStatsModel, PlayerModel.id == PlayerStatsModel.player_id) \
            .filter(PlayerModel.tournament_id == self.tournament_id) \
            .filter(PlayerStatsModel.final_rank.isnot(None)) \
            .order_by(PlayerStatsModel.final_rank.asc()) \
            .all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Konečné pořadí"

        # Styly
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        align_center = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        border_thin = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='D1D5DB'),
            right=openpyxl.styles.Side(style='thin', color='D1D5DB'),
            top=openpyxl.styles.Side(style='thin', color='D1D5DB'),
            bottom=openpyxl.styles.Side(style='thin', color='D1D5DB')
        )

        # Nadpis v Excelu
        ws.merge_cells("A1:B1")
        ws["A1"] = f"Výsledky turnaje: {self.tournament.name}"
        ws["A1"].font = openpyxl.styles.Font(size=14, bold=True)
        ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")

        # Hlavička tabulky
        headers = ["Pořadí", "Hráč"]
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_thin

        # Zápis dat
        for row_idx, (player, rank) in enumerate(results_data, start=4):
            c1 = ws.cell(row=row_idx, column=1, value=f"{rank}.")
            c2 = ws.cell(row=row_idx, column=2, value=player.name)

            c1.alignment = align_center
            c1.border = border_thin
            c2.border = border_thin

        # Šířka sloupců
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_length + 5, 15)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return file_stream